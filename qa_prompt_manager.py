import hashlib
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

import numpy as np
from pythainlp.corpus import thai_stopwords
from pythainlp.tokenize import word_tokenize


class QAPromptManager:
    """Manage QA prompt data from Excel and provide semantic + keyword matching."""

    COLUMN_SOURCE_DOCUMENT = "ภาคและเล่มเอกสาร"
    COLUMN_CATEGORY = "หมวดหมู่"
    COLUMN_QUESTION = "คำถาม"

    def __init__(
        self,
        excel_path: str = "./data/QA.xlsx",
        json_path: str = "./data/qa_prompt.json",
        embedding_model: Optional[Any] = None,
        match_threshold: float = 0.50,
    ):
        self.excel_path = Path(excel_path)
        self.json_path = Path(json_path)
        self.embedding_model = embedding_model
        self.match_threshold = match_threshold

        self._stopwords = set(thai_stopwords())
        self._items: List[Dict[str, Any]] = []
        self._answered_items: List[Dict[str, Any]] = []
        self._answered_embeddings: Optional[np.ndarray] = None
        self._answered_tokens: List[set[str]] = []
        self._index_ready = False

    def sync_from_excel(self, force: bool = False, preserve_answers: bool = True) -> Dict[str, Any]:
        """Sync Excel -> JSON when needed, preserving non-empty answers from existing JSON."""
        should_sync = force or self._should_sync()
        if should_sync:
            items = self._build_items_from_excel(preserve_answers=preserve_answers)
            self._write_items(items)
        else:
            items = self.load_items()

        self._set_items(items)
        return {
            "updated": should_sync,
            "count": len(items),
            "excel_path": str(self.excel_path),
            "json_path": str(self.json_path),
        }

    def load_items(self) -> List[Dict[str, Any]]:
        """Load QA items from JSON."""
        if not self.json_path.exists():
            return []

        with self.json_path.open("r", encoding="utf-8") as f:
            data = json.load(f)

        if isinstance(data, list):
            return [self._normalize_item(item) for item in data if isinstance(item, dict)]
        if isinstance(data, dict) and isinstance(data.get("items"), list):
            return [self._normalize_item(item) for item in data["items"] if isinstance(item, dict)]
        return []

    def find_best_match(self, query: str, expanded_query: Optional[str] = None) -> Optional[Dict[str, Any]]:
        """
        Find best answered QA candidate using hybrid score:
        combined_score = 0.7 * semantic + 0.3 * keyword
        """
        if not query or not query.strip():
            return None
        self._ensure_index()
        if not self._answered_items:
            return None
        if self.embedding_model is None:
            return None

        variants = [query.strip()]
        if expanded_query and expanded_query.strip() and expanded_query.strip() != variants[0]:
            variants.append(expanded_query.strip())

        variant_embeddings = self.embedding_model.encode(variants, normalize_embeddings=True)
        variant_tokens = [self._tokenize(text) for text in variants]

        best_match: Optional[Dict[str, Any]] = None

        for idx, item in enumerate(self._answered_items):
            item_best_score = 0.0
            item_best_semantic = 0.0
            item_best_keyword = 0.0
            item_best_variant = variants[0]

            for variant_idx, variant_text in enumerate(variants):
                semantic_score = float(np.dot(self._answered_embeddings[idx], variant_embeddings[variant_idx]))
                keyword_score = self._keyword_overlap(variant_tokens[variant_idx], self._answered_tokens[idx])
                combined_score = (0.7 * semantic_score) + (0.3 * keyword_score)

                if combined_score > item_best_score:
                    item_best_score = combined_score
                    item_best_semantic = semantic_score
                    item_best_keyword = keyword_score
                    item_best_variant = variant_text

            if item_best_score < self.match_threshold:
                continue

            candidate = dict(item)
            candidate["semantic_score"] = item_best_semantic
            candidate["keyword_score"] = item_best_keyword
            candidate["combined_score"] = item_best_score
            candidate["matched_variant"] = item_best_variant

            if best_match is None or candidate["combined_score"] > best_match["combined_score"]:
                best_match = candidate

        return best_match

    def _should_sync(self) -> bool:
        """Return True when JSON missing or Excel has newer mtime."""
        if not self.excel_path.exists():
            return False
        if not self.json_path.exists():
            return True
        return self.excel_path.stat().st_mtime > self.json_path.stat().st_mtime

    def _build_items_from_excel(self, preserve_answers: bool = True) -> List[Dict[str, Any]]:
        items = self._parse_excel_items()
        if not preserve_answers:
            return items

        existing_items = self.load_items()
        existing_by_id = {item["id"]: item for item in existing_items if item.get("id")}

        for item in items:
            old = existing_by_id.get(item["id"])
            if not old:
                continue
            old_answer = str(old.get("answer", "")).strip()
            if old_answer:
                item["answer"] = old.get("answer", "")
                item["answer_updated_at"] = old.get("answer_updated_at")

        return items

    def _parse_excel_items(self) -> List[Dict[str, Any]]:
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required to read QA.xlsx. Please install with: pip install openpyxl"
            ) from exc

        workbook = load_workbook(filename=self.excel_path, data_only=True, read_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)

        headers = next(rows, None)
        if headers is None:
            workbook.close()
            return []

        header_index = {str(value).strip(): idx for idx, value in enumerate(headers) if value is not None}
        required_cols = [
            self.COLUMN_SOURCE_DOCUMENT,
            self.COLUMN_CATEGORY,
            self.COLUMN_QUESTION,
        ]

        missing = [col for col in required_cols if col not in header_index]
        if missing:
            workbook.close()
            raise ValueError(f"Missing required columns in Excel: {missing}")

        source_idx = header_index[self.COLUMN_SOURCE_DOCUMENT]
        category_idx = header_index[self.COLUMN_CATEGORY]
        question_idx = header_index[self.COLUMN_QUESTION]

        current_source = ""
        current_category = ""
        items: List[Dict[str, Any]] = []

        for row in rows:
            source_value = self._clean_cell(row[source_idx] if source_idx < len(row) else "")
            category_value = self._clean_cell(row[category_idx] if category_idx < len(row) else "")
            question_value = self._clean_cell(row[question_idx] if question_idx < len(row) else "")

            if source_value:
                current_source = source_value
            if category_value:
                current_category = category_value

            if not question_value:
                continue

            normalized_question = self._normalize_question(question_value)
            item_id = hashlib.sha1(normalized_question.encode("utf-8")).hexdigest()
            items.append(
                {
                    "id": item_id,
                    "source_document": current_source,
                    "category": current_category,
                    "question": question_value,
                    "answer": "",
                    "answer_updated_at": None,
                }
            )

        workbook.close()
        return items

    def _write_items(self, items: List[Dict[str, Any]]) -> None:
        self.json_path.parent.mkdir(parents=True, exist_ok=True)
        with self.json_path.open("w", encoding="utf-8") as f:
            json.dump(items, f, ensure_ascii=False, indent=2)

    def _set_items(self, items: List[Dict[str, Any]]) -> None:
        self._items = items
        self._index_ready = False

    def _ensure_index(self) -> None:
        if self._index_ready:
            return
        if self.embedding_model is None:
            self._answered_items = []
            self._answered_embeddings = None
            self._answered_tokens = []
            self._index_ready = True
            return
        if not self._items:
            self._items = self.load_items()

        self._answered_items = [item for item in self._items if str(item.get("answer", "")).strip()]
        if not self._answered_items:
            self._answered_embeddings = None
            self._answered_tokens = []
            self._index_ready = True
            return

        questions = [item["question"] for item in self._answered_items]
        self._answered_embeddings = self.embedding_model.encode(questions, normalize_embeddings=True)
        self._answered_tokens = [self._tokenize(question) for question in questions]
        self._index_ready = True

    def _tokenize(self, text: str) -> set[str]:
        words = word_tokenize(text, engine="newmm")
        filtered = set()
        for word in words:
            normalized = word.strip().lower()
            if len(normalized) < 2:
                continue
            if normalized in self._stopwords:
                continue
            if normalized.isdigit():
                continue
            filtered.add(normalized)
        return filtered

    @staticmethod
    def _keyword_overlap(query_tokens: set[str], question_tokens: set[str]) -> float:
        if not query_tokens:
            return 0.0
        return len(query_tokens.intersection(question_tokens)) / len(query_tokens)

    @staticmethod
    def _clean_cell(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _normalize_question(question: str) -> str:
        cleaned = re.sub(r"\s+", " ", question.strip())
        return cleaned.lower()

    @staticmethod
    def _normalize_item(item: Dict[str, Any]) -> Dict[str, Any]:
        question = str(item.get("question", "")).strip()
        normalized_question = QAPromptManager._normalize_question(question)
        item_id = item.get("id") or hashlib.sha1(normalized_question.encode("utf-8")).hexdigest()
        return {
            "id": str(item_id),
            "source_document": str(item.get("source_document", "")).strip(),
            "category": str(item.get("category", "")).strip(),
            "question": question,
            "answer": item.get("answer", "") if item.get("answer") is not None else "",
            "answer_updated_at": item.get("answer_updated_at"),
        }
